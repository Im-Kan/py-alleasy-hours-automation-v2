from openpyxl import load_workbook
from glob import glob
import api.automation.time_explorer as time_explorer
import api.automation.alleasy as alleasy
import api.automation.timetracker_api as timetracker_api
import calendar
import base64
import os
from io import BytesIO


def _build_alleasy(calendar_te, hour_hashmap, mes, ano, nome):
    """Shared logic: fills the AllEasy template and returns base64."""
    _, dias_no_mes = calendar.monthrange(ano, mes)
    alleasy_file = glob("./api/automation/alleasy_model_excel/*.xlsx")[0]
    wk_alleasy = load_workbook(alleasy_file)
    ws_alleasy = wk_alleasy.active

    alleasy.set_data(ws_alleasy, mes, ano, dias_no_mes)
    alleasy.set_horas(ws_alleasy, calendar_te, dias_no_mes, hour_hashmap)
    alleasy.set_nome(ws_alleasy, nome)

    bytes_io = BytesIO()
    wk_alleasy.save(bytes_io)
    return base64.b64encode(bytes_io.getvalue()).decode('utf-8')


def generateAlleasyExcel(data):
    """Original flow: user uploads Time Explorer xlsx."""
    mes = int(data["date"].split("/")[0])
    ano = int(data["date"].split("/")[1])

    decrypted = base64.b64decode(data["base64"])
    devops_file = BytesIO(decrypted)
    wk_devops = load_workbook(devops_file)
    ws_devops = wk_devops.active

    (calendar_te, hour_hashmap) = time_explorer.get_calendar_hashmap(ws_devops)
    return _build_alleasy(calendar_te, hour_hashmap, mes, ano, data["name"])


def generateAlleasyFromAPI(data):
    """New flow: fetches data directly from 7pace Reporting API."""
    mes = int(data["date"].split("/")[0])
    ano = int(data["date"].split("/")[1])

    api_url = os.environ.get('TIMETRACKER_API_URL', '')
    api_token = os.environ.get('TIMETRACKER_API_TOKEN', '')

    # Fallback: read from .env file if system env vars are missing
    if not api_url or not api_token:
        env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), '.env')
        if os.path.exists(env_path):
            with open(env_path, 'r') as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    key, _, value = line.partition('=')
                    if key == 'TIMETRACKER_API_URL' and not api_url:
                        api_url = value
                    elif key == 'TIMETRACKER_API_TOKEN' and not api_token:
                        api_token = value

    if not api_url or not api_token:
        raise ValueError("TIMETRACKER_API_URL and TIMETRACKER_API_TOKEN not found in env vars or .env file")

    api = timetracker_api.TimetrackerAPI(api_url, api_token)
    worklogs = api.get_worklogs_work_items(ano, mes)
    rows = api.to_rows(worklogs)

    (calendar_te, hour_hashmap) = time_explorer.get_calendar_hashmap_from_api(rows)
    return _build_alleasy(calendar_te, hour_hashmap, mes, ano, data["name"])
