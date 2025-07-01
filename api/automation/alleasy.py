from openpyxl.styles import PatternFill
from datetime import time, timedelta


def set_data(ws, mes, ano, dias_no_mes):
    ws['B4'].value = "01/"+mes.__str__()+"/"+ano.__str__()
    ws['E4'].value = dias_no_mes.__str__()+"/"+mes.__str__()+"/"+ano.__str__()


def set_horas(ws, calendar_hashmap, dias_do_mes, hour_hashmap):
    n = 10
    for dia in range(1, dias_do_mes+1):
        n = n+1
        obs = calendar_hashmap.get(dia)
        if obs is None:
            fill_empty(ws, n, obs)
            continue
        fill_work(ws, n, obs,hour_hashmap,dia)
        print(n, ' ', dia, "=>", obs)


def set_nome(ws, nome):
    ws['B2'].value = nome

def time_to_timedelta(t):
    return timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)

def format_timedelta(td):
    total_seconds = int(td.total_seconds())
    hours = total_seconds // 3600
    if(hours >23):
        hours = hours-24
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return time(hours,minutes,seconds)

def fill_work(ws, n, obs,hour_hashmap,dia):
    fill_white(ws, n, obs)
    hora = time(8, 0, 0)
    hora_total = hour_hashmap[dia]
    hora_extra = 0
    hora_extra_time = time(0,0,0)
    if hora_total > 8:
        hora_extra = hora_total - 8 
        hour_extra   = int(hora_extra)
        minute_extra = int((hora_extra - hour_extra) * 60)
        sec_extra    = int(((hora_extra - hour_extra) * 60 - minute_extra) * 60)
        hora_extra_time = time(hour_extra, minute_extra, sec_extra)
    hora_total_time = format_timedelta(time_to_timedelta(hora) + time_to_timedelta(hora_extra_time))
    ws.row_dimensions[n].height = 30
    ws['P'+n.__str__()].value = obs
    ws['E'+n.__str__()].value = "09:00:00"
    ws['F'+n.__str__()].value = "12:00:00"
    ws['G'+n.__str__()].value = "13:00:00"
    ws['H'+n.__str__()].value = "18:00:00"
    ws['K'+n.__str__()].value = hora
    if(hora_extra>0):
        ws['I'+n.__str__()].value = "18:00:00"
        h = timedelta(hours=18,minutes=0,seconds=0) + time_to_timedelta(hora_extra_time)
        ws['J'+n.__str__()].value = format_timedelta(h)
        ws['L'+n.__str__()].value = hora_extra_time
    
    ws['M'+n.__str__()].value = hora_total_time


def fill_empty(ws, n, obs):
    gray = "C0C0C0"
    empty_filler = PatternFill(start_color=gray, end_color=gray, fill_type="solid")

    ws['A'+n.__str__()].fill = empty_filler
    ws['B'+n.__str__()].fill = empty_filler
    ws['C'+n.__str__()].fill = empty_filler
    ws['D'+n.__str__()].fill = empty_filler
    ws['E'+n.__str__()].fill = empty_filler
    ws['F'+n.__str__()].fill = empty_filler
    ws['G'+n.__str__()].fill = empty_filler
    ws['H'+n.__str__()].fill = empty_filler
    ws['I'+n.__str__()].fill = empty_filler
    ws['J'+n.__str__()].fill = empty_filler
    ws['K'+n.__str__()].fill = empty_filler
    ws['L'+n.__str__()].fill = empty_filler
    ws['M'+n.__str__()].fill = empty_filler
    ws['P'+n.__str__()].fill = empty_filler


def fill_white(ws, n, obs):
    color = "ffffff"
    empty_filler = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws['A'+n.__str__()].fill = empty_filler
    ws['B'+n.__str__()].fill = empty_filler
    ws['C'+n.__str__()].fill = empty_filler
    ws['D'+n.__str__()].fill = empty_filler
    ws['E'+n.__str__()].fill = empty_filler
    ws['F'+n.__str__()].fill = empty_filler
    ws['G'+n.__str__()].fill = empty_filler
    ws['H'+n.__str__()].fill = empty_filler
    ws['I'+n.__str__()].fill = empty_filler
    ws['J'+n.__str__()].fill = empty_filler
    ws['K'+n.__str__()].fill = empty_filler
    ws['L'+n.__str__()].fill = empty_filler
    ws['M'+n.__str__()].fill = empty_filler
    ws['P'+n.__str__()].fill = empty_filler
    # while True:
    #     dia = ws['B'+n.__str__()].value.__str__()
    #     if dia is None:
    #         break
    #
    #     print("diaa", dia)
    #     obs = calendar_hashmap[dia]
    #     ws['P'+n.__str__()].value = obs
    #     print(dia, "=>", obs)

